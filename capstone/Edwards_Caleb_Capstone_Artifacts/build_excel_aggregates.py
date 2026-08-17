# build_excel_aggregates.py
# Builds CAPSTONE_EXCEL_AGGREGATES.xlsx.
# WHAT: three aggregate tabs (prescriber / specialty / generic drug) plus a
#       reconciliation tab that ties every tab back to the cleaned dataset.
# WHY:  Stage 0 verified CAPSTONE_AGGREGATES.xlsx PASS, so the group-level totals
#       are copied from it rather than re-derived. The two per-unit metrics and the
#       reconciliation are written as live Excel formulas (SUM, IFERROR, absolute
#       references, ROUND, IF, COUNTA) so every number on the sheet is recomputable
#       in Excel itself. Techniques limited to CALIBRATION.md section 3 (Excel taught list).

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = 'CAPSTONE_AGGREGATES.xlsx'
OUT = 'CAPSTONE_EXCEL_AGGREGATES.xlsx'

# Expected totals, computed once from outputs\part_d_co_clean.csv (see PROGRESS.md Stage 1).
EXP = {'rows': 390473, 'cost': 2737455388.61, 'clms': 16573710, 'fills': 33119594.5,
       'npi': 19390, 'spec': 97, 'gnrc': 1177}

src = openpyxl.load_workbook(SRC, data_only=True)
out = openpyxl.Workbook()
out.remove(out.active)

HDR_FILL = PatternFill('solid', fgColor='DDEBF7')
TOT_FILL = PatternFill('solid', fgColor='FFF2CC')
BOLD = Font(bold=True)

# Tabs to carry over, with the column index (1-based, in the SOURCE sheet) of the
# three additive columns we re-total: cost, claims, 30-day fills.
TABS = [
    ('By_Prescriber',   'Prscrbr_NPI',              19390),
    ('By_Specialty',    'Prscrbr_Type (specialty)', 97),
    ('By_Generic_Drug', 'Gnrc_Name',                1177),
]

built = {}

for tab, keydesc, ngroups in TABS:
    ws_src = src[tab]
    rows = list(ws_src.iter_rows(values_only=True))
    provenance = rows[0][0]
    header = list(rows[2])
    data = [list(r) for r in rows[3:] if r[0] is not None]

    # Locate the additive columns by header name, so the code does not depend on order.
    i_cost  = header.index('Total Drug Cost')
    i_clms  = header.index('Total Claims')
    i_fills = header.index('Total 30-Day Fills')
    i_cpc   = header.index('Cost per Claim')
    i_cum   = header.index('Cumulative % of Total Cost')

    # New header: insert "Cost per 30-Day Fill" immediately after "Cost per Claim".
    new_header = header[:i_cpc+1] + ['Cost per 30-Day Fill'] + header[i_cpc+1:]

    ws = out.create_sheet(tab)
    ws['A1'] = provenance
    ws['A1'].font = Font(italic=True, size=9)

    # Row 2 = a visible grand-total row. Every cumulative-% formula below anchors to it
    # with an absolute reference, and the Reconciliation tab reads it directly.
    ws['A2'] = 'GRAND TOTAL (SUM of the column below)'
    ws['A2'].font = BOLD

    for c, h in enumerate(new_header, start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = BOLD
        cell.fill = HDR_FILL
        cell.alignment = Alignment(wrap_text=True, vertical='bottom')

    first, last = 4, 3 + len(data)

    # Column letters in the NEW layout (one column inserted after Cost per Claim).
    def L(src_idx):
        shift = 1 if src_idx > i_cpc else 0
        return get_column_letter(src_idx + 1 + shift)
    cCost, cClms, cFills = L(i_cost), L(i_clms), L(i_fills)
    cCpc = get_column_letter(i_cpc + 1)
    cCpf = get_column_letter(i_cpc + 2)      # the newly inserted column
    cCum = L(i_cum)

    for r_off, row in enumerate(data):
        r = first + r_off
        vals = row[:i_cpc+1] + [None] + row[i_cpc+1:]
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        # Cost per claim = total cost / total claims. IFERROR guards a zero denominator.
        ws[f'{cCpc}{r}'] = f'=IFERROR({cCost}{r}/{cClms}{r},"")'
        # Cost per 30-day fill = total cost / total 30-day fills (dose-standardised unit cost).
        ws[f'{cCpf}{r}'] = f'=IFERROR({cCost}{r}/{cFills}{r},"")'
        # Running share of statewide spend; $-anchored to the grand total in row 2.
        if r == first:
            ws[f'{cCum}{r}'] = f'={cCost}{r}/${cCost}$2*100'
        else:
            ws[f'{cCum}{r}'] = f'={cCum}{r-1}+{cCost}{r}/${cCost}$2*100'

    # Grand-total formulas in row 2 for the three additive columns.
    for col in (cCost, cClms, cFills):
        c = ws[f'{col}2']
        c.value = f'=SUM({col}{first}:{col}{last})'
        c.font = BOLD
        c.fill = TOT_FILL

    for col, w in (('A', 24), (cCost, 18), (cClms, 14), (cFills, 18), (cCpc, 14), (cCpf, 16), (cCum, 16)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A4'

    for r in range(first, last+1):
        ws[f'{cCost}{r}'].number_format = '#,##0.00'
        ws[f'{cClms}{r}'].number_format = '#,##0'
        ws[f'{cFills}{r}'].number_format = '#,##0.0'
        ws[f'{cCpc}{r}'].number_format = '#,##0.00'
        ws[f'{cCpf}{r}'].number_format = '#,##0.00'
        ws[f'{cCum}{r}'].number_format = '0.00'
    ws[f'{cCost}2'].number_format = '#,##0.00'
    ws[f'{cClms}2'].number_format = '#,##0'
    ws[f'{cFills}2'].number_format = '#,##0.0'

    built[tab] = dict(cost=cCost, clms=cClms, fills=cFills, first=first, last=last,
                      ngroups=ngroups, keydesc=keydesc, rowcount=len(data))

# ---------------- Reconciliation tab ----------------
# Each aggregate tab is a complete partition of the 390,473 cleaned rows, so each tab's
# column totals must equal the cleaned-file totals exactly. Delta and Status are formulas
# so the check re-evaluates whenever the workbook is opened.
rec = out.create_sheet('Reconciliation', 0)
rec['A1'] = 'RECONCILIATION -- every aggregate tab tied back to outputs\\part_d_co_clean.csv'
rec['A1'].font = Font(bold=True, size=12)
rec['A2'] = ('Expected values were computed once from the cleaned file (390,473 rows). '
             'Each aggregate tab is a complete partition of those rows, so each tab must reproduce every total exactly.')
rec['A2'].font = Font(italic=True, size=9)

rec['A4'] = 'Cleaned dataset -- reference totals'
rec['A4'].font = BOLD
ref = [('Rows in cleaned file', EXP['rows'], '#,##0'),
       ('Total drug cost ($)', EXP['cost'], '#,##0.00'),
       ('Total claims', EXP['clms'], '#,##0'),
       ('Total 30-day fills', EXP['fills'], '#,##0.0'),
       ('Distinct prescribers (NPI)', EXP['npi'], '#,##0'),
       ('Distinct specialties', EXP['spec'], '#,##0'),
       ('Distinct generic drugs', EXP['gnrc'], '#,##0')]
for i, (lab, val, fmt) in enumerate(ref):
    rec.cell(row=5+i, column=1, value=lab)
    c = rec.cell(row=5+i, column=2, value=val); c.number_format = fmt

start = 5 + len(ref) + 2
rec.cell(row=start-1, column=1, value='Tie-out checks').font = BOLD
heads = ['#', 'Check', 'Tab', 'Expected', 'Workbook value (live formula)', 'Delta', 'Status']
for c, h in enumerate(heads, start=1):
    cell = rec.cell(row=start, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL

checks = []
for tab in ['By_Prescriber', 'By_Specialty', 'By_Generic_Drug']:
    b = built[tab]
    checks += [
        ('Total drug cost ($)',  tab, EXP['cost'],  f"='{tab}'!{b['cost']}2",  '#,##0.00', 2),
        ('Total claims',         tab, EXP['clms'],  f"='{tab}'!{b['clms']}2",  '#,##0',    0),
        ('Total 30-day fills',   tab, EXP['fills'], f"='{tab}'!{b['fills']}2", '#,##0.0',  1),
        ('Group count (rows)',   tab, b['ngroups'], f"=COUNTA('{tab}'!A{b['first']}:A{b['last']})", '#,##0', 0),
    ]

r = start + 1
for n, (lab, tab, exp, formula, fmt, dp) in enumerate(checks, start=1):
    rec.cell(row=r, column=1, value=n)
    rec.cell(row=r, column=2, value=lab)
    rec.cell(row=r, column=3, value=tab)
    c = rec.cell(row=r, column=4, value=exp); c.number_format = fmt
    c = rec.cell(row=r, column=5, value=formula); c.number_format = fmt
    c = rec.cell(row=r, column=6, value=f'=ROUND(E{r}-D{r},{dp})'); c.number_format = fmt
    rec.cell(row=r, column=7, value=f'=IF(F{r}=0,"TIE","CHECK")')
    r += 1

rec.cell(row=r+1, column=1,
         value='Note: source row count (390,473) is not summable from an aggregate tab -- each tab collapses those rows into groups. '
               'It is verified instead by the group counts above matching the distinct-entity counts of the cleaned file.').font = Font(italic=True, size=9)

for col, w in (('A', 6), ('B', 26), ('C', 18), ('D', 20), ('E', 26), ('F', 14), ('G', 10)):
    rec.column_dimensions[col].width = w
rec.column_dimensions['A'].width = 6

# ---------------- Index ----------------
idx = out.create_sheet('0_Index', 0)
idx['A1'] = 'CAPSTONE_EXCEL_AGGREGATES.xlsx'
idx['A1'].font = Font(bold=True, size=12)
idx['A2'] = ('Source: outputs\\part_d_co_clean.csv (390,473 rows, Colorado only). Group totals carried over from '
             'CAPSTONE_AGGREGATES.xlsx, verified PASS in VERIFICATION.md check 2. Per-unit metrics and all '
             'reconciliation values are live Excel formulas.')
idx['A2'].font = Font(italic=True, size=9)
idx['A2'].alignment = Alignment(wrap_text=True)
for c, h in enumerate(['Sheet', 'Grouped by', 'Groups (rows)', 'Metrics'], start=1):
    cell = idx.cell(row=4, column=c, value=h); cell.font = BOLD; cell.fill = HDR_FILL
metrics = 'Total drug cost, total claims, total 30-day fills, cost per claim, cost per 30-day fill, cumulative % of total cost'
ir = 5
idx.cell(row=ir, column=1, value='Reconciliation'); idx.cell(row=ir, column=2, value='--')
idx.cell(row=ir, column=3, value=len(checks)); idx.cell(row=ir, column=4, value='Tie-out of every tab to the cleaned dataset totals')
ir += 1
for tab, keydesc, n in TABS:
    idx.cell(row=ir, column=1, value=tab); idx.cell(row=ir, column=2, value=keydesc)
    idx.cell(row=ir, column=3, value=n); idx.cell(row=ir, column=4, value=metrics)
    ir += 1
for col, w in (('A', 20), ('B', 26), ('C', 14), ('D', 70)):
    idx.column_dimensions[col].width = w

out.save(OUT)
print('wrote', OUT)
